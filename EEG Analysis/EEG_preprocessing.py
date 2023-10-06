import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

Participant_list = ["H1", "H2", "H3", "H4", "H5", "H6"]

df_pre_list = []
df_post_list = []

def ModifyExcelSheet(i,state) :

    wb = load_workbook(filename="EEG Data/" + i + "_EC_" + state + "FrequencysPowerandPeak_Combined.xlsx")
    ws = wb.active
    ws.move_range("A1:I66", rows=1, cols=0, translate=True)

    index = 1

    for iter in range(1,67): 

        ws[get_column_letter(index) + str(1)] = "Channel_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Delta_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Delta_frequency_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Theta_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Theta_frequency_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Alpha_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Alpha_frequency_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Beta_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Beta_frequency_" + str(iter)
        index += 1

    ws[get_column_letter(index)+str(1)] = "Output"
    
    if(state=="pre"):

        ws[get_column_letter(index) + str(2)] = 0

    if(state=="post"):

        ws[get_column_letter(index) + str(2)] = 1

    for iter in range(2,68):

        ws.move_range("A" + str(iter) + ":I" + str(iter), rows=-1*(iter-2), cols=9*(iter-2), translate=True)

    wb.save(filename="EEG Data Modified/" + i + "_EC_" + state + "FrequencysPowerandPeak_Combined.xlsx")

for i in Participant_list :

    ModifyExcelSheet(i,"pre")
    ModifyExcelSheet(i,"post")

    df_pre = pd.read_excel("EEG Data Modified/" + i + "_EC_" + "preFrequencysPowerandPeak_Combined.xlsx")
    df_post = pd.read_excel("EEG Data Modified/" + i + "_EC_" + "postFrequencysPowerandPeak_Combined.xlsx")

    df_pre_list.append(df_pre)
    df_post_list.append(df_post)

Pre_list_merged_df = pd.concat(df_pre_list, ignore_index=True, sort=False)
Post_list_merged_df = pd.concat(df_post_list, ignore_index=True, sort=False)

channel_list = []

for i in range(1,67):
    channel_list.append("Channel_" + str(i))

Pre_list_merged_df = Pre_list_merged_df.drop(channel_list, axis = 1)
Post_list_merged_df = Post_list_merged_df.drop(channel_list, axis = 1)

Final_df = pd.concat([Pre_list_merged_df,Post_list_merged_df], ignore_index=True, sort=False)
Final_df.to_excel("Final_df.xlsx")