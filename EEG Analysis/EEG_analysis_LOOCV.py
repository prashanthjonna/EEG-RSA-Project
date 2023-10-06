import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from numpy import mean
from numpy import absolute
from numpy import sqrt
from sklearn import svm
from sklearn.svm import SVC
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn.model_selection import cross_val_score
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import LeaveOneOut
from xgboost import XGBClassifier
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis as LDA
from flaml import AutoML


def cap_data(df):
    for col in df.columns:
        print("capping the ",col)
        if (((df[col].dtype)=='float64') | ((df[col].dtype)=='int64')):
            percentiles = df[col].quantile([0.20,0.80]).values
            df[col][df[col] <= percentiles[0]] = percentiles[0]
            df[col][df[col] >= percentiles[1]] = percentiles[1]
        else:
            df[col]=df[col]
    return df


def RunModels(X_train,y_train,X_test,y_test,DR_tech):

    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)

    if(DR_tech == 1):
        pca = PCA(n_components = 0.995)
        X_train = pca.fit_transform(X_train)
        X_test = pca.transform(X_test)

    if(DR_tech == 2):
        lda = LDA(n_components = 1)
        X_train = lda.fit_transform(X_train,y_train)
        X_test = lda.transform(X_test)

    LR_classifier = LogisticRegression(random_state = 0)
    LR_classifier.fit(X_train, y_train)
    print("Logistic regression score : ", LR_classifier.score(X_test,y_test))

    XGB_classifier = XGBClassifier()
    XGB_classifier.fit(X_train, y_train)
    print("XG Boost score : ", XGB_classifier.score(X_test,y_test))

    KNN_classifier = KNeighborsClassifier(n_neighbors = 3, metric = 'minkowski', p = 2)
    KNN_classifier.fit(X_train, y_train)
    print("KNN score : ", KNN_classifier.score(X_test,y_test))

    SVM_classifier = SVC(kernel = 'linear', random_state = 0)
    SVM_classifier.fit(X_train, y_train)
    print("SVM score : ", SVM_classifier.score(X_test,y_test))

    KSVM_classifier = SVC(kernel = 'rbf', random_state = 0)
    KSVM_classifier.fit(X_train, y_train)
    print("Kernel SVM score : ", KSVM_classifier.score(X_test,y_test))

    NB_classifier = GaussianNB()
    NB_classifier.fit(X_train, y_train)
    print("Naive bayes score : ", NB_classifier.score(X_test,y_test))

    DTC_classifier = DecisionTreeClassifier(criterion = 'entropy', random_state = 0)
    DTC_classifier.fit(X_train, y_train)
    print("Decision Tree score : ", DTC_classifier.score(X_test,y_test))

    RFC_classifier = RandomForestClassifier(n_estimators = 10, criterion = 'entropy', random_state = 0)
    RFC_classifier.fit(X_train, y_train)
    print("Random Forest Classifier score : ", RFC_classifier.score(X_test,y_test))
    



Participant_list = ["H1", "H2", "H3"]

df_pre_list = []
df_post_list = []


def ModifyExcelSheet(i,state) :

    wb = load_workbook(filename="EEG Data/" + i + "_EC_" + state + "FrequencysPowerandPeak.xlsx")
    ws = wb.active
    ws.move_range("A1:I66", rows=1, cols=0, translate=True)

    index = 1

    for iter in range(1,67): 

        ws[get_column_letter(index) + str(1)] = "Channel_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Delta_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Delta_frequency_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Theta_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Theta_frequency_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Alpha_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Alpha_frequency_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Beta_value_" + str(iter)
        index += 1
        ws[get_column_letter(index) + str(1)] = "Beta_frequency_" + str(iter)
        index += 1

    ws[get_column_letter(index)+str(1)] = "Output"
    
    if(state=="pre"):

        ws[get_column_letter(index) + str(2)] = 0

    if(state=="post"):

        ws[get_column_letter(index) + str(2)] = 1

    for iter in range(2,68):

        ws.move_range("A" + str(iter) + ":I" + str(iter), rows=-1*(iter-2), cols=9*(iter-2), translate=True)

    wb.save(filename="EEG Data Modified/" + i + "_EC_" + state + "FrequencysPowerandPeak.xlsx")


for i in Participant_list :

    ModifyExcelSheet(i,"pre")
    ModifyExcelSheet(i,"post")

    df_pre = pd.read_excel("EEG Data Modified/" + i + "_EC_" + "preFrequencysPowerandPeak.xlsx")
    df_post = pd.read_excel("EEG Data Modified/" + i + "_EC_" + "postFrequencysPowerandPeak.xlsx")

    df_pre_list.append(df_pre)
    df_post_list.append(df_post)


Pre_list_merged_df = pd.concat(df_pre_list, ignore_index=True, sort=False)
Post_list_merged_df = pd.concat(df_post_list, ignore_index=True, sort=False)

channel_list = []

for i in range(1,67):
    channel_list.append("Channel_" + str(i))

Pre_list_merged_df = Pre_list_merged_df.drop(channel_list, axis = 1)
Post_list_merged_df = Post_list_merged_df.drop(channel_list, axis = 1)

Final_df = pd.concat([Pre_list_merged_df,Post_list_merged_df], ignore_index=True, sort=False)

# print(Final_df)

X = Final_df.iloc[:,:-1].values
y = Final_df.iloc[:, -1].values
X = pd.DataFrame(X)
y = pd.DataFrame(y)
#X = cap_data(X)

DR_algo = 3

X_train = X.drop(labels = [0,3], axis = 0)
X_test = X.drop(labels = [1,2,4,5], axis = 0)

y_train = y.drop(labels = [0,3], axis = 0)
y_test = y.drop(labels = [1,2,4,5], axis = 0)

print("\nLeaving out participant H1 : \n")
RunModels(X_train,y_train.values.ravel(),X_test,y_test.values.ravel(),DR_algo)

X_train = X.drop(labels = [1,4], axis = 0)
X_test = X.drop(labels = [0,2,3,5], axis = 0)

y_train = y.drop(labels = [1,4], axis = 0)
y_test = y.drop(labels = [0,2,3,5], axis = 0)

print("\nLeaving out participant H2 : \n")
RunModels(X_train,y_train.values.ravel(),X_test,y_test.values.ravel(),DR_algo)

X_train = X.drop(labels = [2,5], axis = 0)
X_test = X.drop(labels = [0,1,3,4], axis = 0)

y_train = y.drop(labels = [2,5], axis = 0)
y_test = y.drop(labels = [0,1,3,4], axis = 0)

print("\Leaving out participant H3 : \n")
RunModels(X_train,y_train.values.ravel(),X_test,y_test.values.ravel(),DR_algo)